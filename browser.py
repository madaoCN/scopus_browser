#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/5/9 10:58 AM
# @Author  : MADAO

from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtWebEngineWidgets import *

from handler.search_list_parser import SearchListParser
from handler.search_detail_parser import SearchDetailParser
from renderDriver import RenderDriver
import sys
import re
from collections import OrderedDict
import queue
import os 

class MainWindow(QMainWindow):

    # noinspection PyUnresolvedReferences
    def __init__(self, url=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 设置窗口标题
        self.setWindowTitle('Scopus Browser')
        # 设置窗口图标
        self.setWindowIcon(QIcon('icons/penguin.png'))
        # 设置窗口大小1200*900
        self.resize(1200, 900)
        self.show()

        # 设置浏览器
        self.browser = QWebEngineView()
        # 指定打开界面的 URL
        self.browser.setUrl(QUrl(url))
        # 添加浏览器到窗口中
        self.setCentralWidget(self.browser)

        ###使用QToolBar创建导航栏，并使用QAction创建按钮
        # 添加导航栏
        navigation_bar = QToolBar('Navigation')
        # 设定图标的大小
        navigation_bar.setIconSize(QSize(16, 16))
        #添加导航栏到窗口中
        self.addToolBar(navigation_bar)

        #QAction类提供了抽象的用户界面action，这些action可以被放置在窗口部件中
        # 添加前进、后退、停止加载和刷新的按钮
        back_button = QAction(QIcon('icons/back.png'), 'Back', self)
        next_button = QAction(QIcon('icons/next.png'), 'Forward', self)
        stop_button = QAction(QIcon('icons/cross.png'), 'stop', self)
        reload_button = QAction(QIcon('icons/renew.png'), 'reload', self)

        back_button.triggered.connect(self.browser.back)
        next_button.triggered.connect(self.browser.forward)
        stop_button.triggered.connect(self.browser.stop)
        reload_button.triggered.connect(self.browser.reload)

        # 将按钮添加到导航栏上
        navigation_bar.addAction(back_button)
        navigation_bar.addAction(next_button)
        navigation_bar.addAction(stop_button)
        navigation_bar.addAction(reload_button)

        #添加URL地址栏
        self.urlbar = QLineEdit()
        # 让地址栏能响应回车按键信号
        self.urlbar.returnPressed.connect(self.navigate_to_url)

        navigation_bar.addSeparator()
        navigation_bar.addWidget(self.urlbar)

        #让浏览器相应url地址的变化
        self.browser.urlChanged.connect(self.renew_urlbar)

        # 添加操作按钮
        action_button = QAction("解析提取引文", self)
        action_button.triggered.connect(self.parse_html)
        navigation_bar.addAction(action_button)

        # 渲染浏览器
        # self.renderWebView = QWebEngineView()
        self.browser.loadFinished.connect(self.render_load_finished)
        # self.browser.loadProgress.connect(self.render_load_progress)

        # 结果
        self.search_result_page_dict = None
        self.init_result_dict()

        # 是否需要解析
        self.parsing_pages_now = False

    def init_result_dict(self):
        self.search_result_page_dict = OrderedDict()

    def parse_html(self, *args):
        # 已经爬取过的return
        # if self.search_result_page_dict[self.browser.page().url()]:
        #     return
    
        self.init_result_dict()
        self.parsing_pages_now = True
        self.browser.page().toHtml(lambda x: self.__parser_html(x))

    def __parser_html(self, html, scrapy_next = True):
        '''
        爬取下一页面
        '''
        parser = SearchListParser(html)
        parser.parse()
        self.search_result_page_dict.update({
            self.browser.page().url():parser
        })
        
        # 加载下一页面
        if scrapy_next and parser.next_pagination_js and len(parser.next_pagination_js):
            # 进入下一个页面
            self.browser.page().runJavaScript(parser.next_pagination_js)

            # self.parsing_pages_now = False
            # # 爬取详情代码
            # self.__craw_detail_page()
        else:
            # 没有要加载的页面
            # self.parsing_pages_now = False
            # print(self.search_result_page_dict)
            # print("抓取成功：=========")
            # print(len(self.search_result_page_dict))
            self.parsing_pages_now = False
            # 爬取详情代码
            self.__craw_detail_page()

    def __craw_detail_page(self):
        '''
        爬取详情页面
        '''
        for url, parser in self.search_result_page_dict.items():
            # SearchListModel
            for list_model in parser.search_list:
                # 加载并解析
                if list_model.title_link:
                    result = RenderDriver.load_detail_page(list_model.title_link)
                    ref_parser = SearchDetailParser(result)
                    ref_parser.parse()
                    # list Model
                    list_model.doi = ref_parser.doi
                    list_model.ref_list = ref_parser.ref_list
                    list_model.ref_model_list = ref_parser.ref_model_list

                    if list_model and list_model.doi.strip() == "10.2308/accr.2000.75.1.93":
                        print(result)
                        exit(0)
            #     break
            # break
        self.__save_2_file()

    def __save_2_file(self):
        file_name = "output.xlsx"
        import urllib.parse
        try:
            parsed_tuple = urllib.parse.parse_qs(self.browser.page().url().toDisplayString())
            file_name = parsed_tuple.get("searchterm1", [""])[0] + ".xlsx"
        except Exception as e:
            pass

        output_file_name, filetype = QFileDialog.getSaveFileName(self, 
                                                                "保存结果", 
                                                                os.path.join(os.path.expanduser("~"), "Downloads", file_name),
                                                                "All Files (*);;Text Files (*.xlsx)")
        if output_file_name == None or len(output_file_name) == 0: return

        import openpyxl
        work_book = openpyxl.Workbook()
        sheet = work_book.create_sheet("源文章列表", index=0)
        # 域
        fileds = ["DOI", "标题", "标题链接", "作者", "年份", "期刊", "期刊编号","期刊链接", "引文"]
        row = 1
        # 写入域
        for index in range(len(fileds)):
            sheet.cell(row=row, column=index+1).value = fileds[index]

        # 写入内容
        for url, parser in self.search_result_page_dict.items():
            for model in parser.search_list:
                row += 1
                items = [
                    model.doi,
                    model.title,
                    model.title_link,
                    "\r\n".join(
                       map(lambda x:"(".join(x) + ")", model.author)
                    ),
                    model.year,
                    model.journal,
                    model.journal_no,
                    model.journal_link,
                ] 
                items.extend(model.ref_list)

                for index in range(len(items)):
                    sheet.cell(row=row, column=index+1).value = items[index]


        sheet = work_book.create_sheet("参考文献", index=1)
        # 域
        fileds = ["原文章DOI", "原文章标题", "原文章年份", "引文DOI", "引文标题", "引文作者", "引文年份", "引文期刊", "引文期刊编号", "引文引用链接", "引文期刊链接", "引文其他"]
        row = 1
        # 写入域
        for index in range(len(fileds)):
            sheet.cell(row=row, column=index+1).value = fileds[index]

        # 写入内容
        for url, parser in self.search_result_page_dict.items():
            for model in parser.search_list:
                for ref_model in model.ref_model_list:
                    row += 1
                    items = [
                        model.doi,
                        model.title,
                        model.year,

                        ref_model.doi,
                        ref_model.title,
                        ref_model.author,
                        ref_model.year,
                        ref_model.journal,
                        ref_model.journal_page,
                        ref_model.ref_link,
                        ref_model.title_link,
                        ref_model.raw
                    ]
                    for index in range(len(items)):
                        sheet.cell(row=row, column=index+1).value = items[index]
                # row += 1
        work_book.save(filename=output_file_name)


    def navigate_to_url(self):
        q = QUrl(self.urlbar.text())
        if q.scheme() == '':
            q.setScheme('http')
        self.browser.setUrl(q)

    def renew_urlbar(self, q):
        # 将当前网页的链接更新到地址栏
        self.urlbar.setText(q.toString())
        self.urlbar.setCursorPosition(0)

    def render_load_finished(self, isOk):
        '''
        渲染成功
        '''
        print("load success: {} for url: {}".format(isOk, self.browser.url()))
        if isOk and self.parsing_pages_now:
            # 渲染成功，进行解析页面
            self.browser.page().toHtml(lambda x: self.__parser_html(x))

    def render_load_progress(self, progress):
        '''
        加载进度
        '''
        print("progress: {} for url: {}".format(progress, self.browser.url()))

    @staticmethod
    def run(url=None):
        # 创建应用
        app = QApplication(sys.argv)
        # 创建主窗口
        window = MainWindow(url=url)
        # 显示窗口
        window.show()
        # 运行应用，并监听事件
        app.exec_()
        

if __name__ == '__main__':
    url = 'https://www.scopus.com/search/form.uri?display=basic&zone=header&origin='
    MainWindow.run(url=url)
    import os
    os.system("pause")